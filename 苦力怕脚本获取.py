import os
import re
import pandas as pd
from datetime import datetime, timedelta
from urllib.parse import urljoin
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class KlpbbsScraper:
    """
    一个面向对象的网页爬虫，用于从Minecraft(我的世界)苦力怕论坛的用户个人主页抓取其所有主题帖信息。
    该爬虫使用Selenium进行浏览器自动化，支持通过Cookie字符串进行登录，能自动翻页，
    并将抓取到的数据（主题名称、发布日期、浏览量、最后回复日期）保存到带样式的Excel报告中。
    """

    def __init__(self, base_url: str, driver_path: str):
        self.base_url = base_url
        self.driver_path = os.path.abspath(driver_path)
        self.report_path = os.getcwd()
        self.driver = None
        self.data = []
        self.now = datetime.now()

    def _log(self, message: str):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{timestamp}] {message}")

    # [已更新] 智能的WebDriver设置方法
    def _setup_driver(self):
        self._log("开始配置WebDriver (已启用混合动力优化)...")
        options = Options()
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-extensions")
        options.add_argument("--start-maximized")
        options.add_argument("log-level=3")
        options.add_argument('--headless=new')
        prefs = {"profile.managed_default_content_settings.images": 2}
        options.add_experimental_option("prefs", prefs)
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        # 判断是否在GitHub Actions环境中运行
        is_github_actions = os.getenv('CI') == 'true'

        try:
            if is_github_actions:
                # 在GitHub Actions中，不指定驱动路径，让Selenium自动查找
                self._log("在GitHub Actions环境中运行，自动检测ChromeDriver...")
                self.driver = webdriver.Chrome(options=options)
            else:
                # 在本地环境中，使用我们手动指定的路径
                self._log(f"在本地环境中运行，使用指定路径: {self.driver_path}")
                service = Service(self.driver_path)
                self.driver = webdriver.Chrome(service=service, options=options)
            
            self._log("WebDriver启动成功。")
        except Exception as e:
            self._log(f"WebDriver启动失败: {e}")
            raise

    def _login_with_cookie(self, cookie_string: str):
        """使用提供的Cookie字符串登录（优化版）"""
        if not cookie_string or "在这里粘贴" in cookie_string:
            self._log("错误：Cookie字符串为空或未被替换。请先在脚本末尾填写Cookie。")
            raise ValueError("未提供有效的Cookie字符串。")

        domain_url = "https://klpbbs.com"
        self._log(f"正在导航到主域 {domain_url} 以便设置Cookie...")
        self.driver.get(domain_url)
        
        self._log("正在清除旧的Cookie...")
        self.driver.delete_all_cookies()

        self._log("正在解析并添加Cookie...")
        cookies_list = cookie_string.split('; ')
        for cookie_item in cookies_list:
            if '=' in cookie_item:
                name, value = cookie_item.split('=', 1)
                cookie_dict = {'name': name.strip(), 'value': value.strip()}
                try:
                    self.driver.add_cookie(cookie_dict)
                except Exception as e:
                    self._log(f"警告：添加Cookie '{name}' 失败。错误: {e}")
        
        self._log("Cookie添加完成。")
        self._log("正在刷新页面以应用Cookie...")
        self.driver.refresh()
        
        WebDriverWait(self.driver, 5).until(lambda d: d.execute_script('return document.readyState') == 'complete')
        self._log("Cookie应用成功，脚本将以登录状态继续。")

    def _take_screenshot(self, filename: str = "error_screenshot.png"):
        path = os.path.join(self.report_path, filename)
        if not self.driver:
            self._log("Driver未初始化，无法截屏。")
            return
        try:
            self.driver.save_screenshot(path)
            self._log(f"已截取屏幕并保存至: {path}")
        except Exception as e:
            self._log(f"截屏失败: {e}")

    def _parse_human_readable_date(self, date_str: str) -> datetime | None:
        date_str = date_str.strip()
        try:
            if "昨天" in date_str:
                time_part = re.search(r'(\d{2}:\d{2})', date_str).group(1)
                date_part = self.now.date() - timedelta(days=1)
                return datetime.strptime(f"{date_part} {time_part}", '%Y-%m-%d %H:%M')
            elif "前天" in date_str:
                time_part = re.search(r'(\d{2}:\d{2})', date_str).group(1)
                date_part = self.now.date() - timedelta(days=2)
                return datetime.strptime(f"{date_part} {time_part}", '%Y-%m-%d %H:%M')
            elif "小时前" in date_str:
                hours = int(re.search(r'(\d+)', date_str).group(1))
                return self.now - timedelta(hours=hours)
            elif "分钟前" in date_str:
                minutes = int(re.search(r'(\d+)', date_str).group(1))
                return self.now - timedelta(minutes=minutes)
            elif "天前" in date_str:
                days = int(re.search(r'(\d+)', date_str).group(1))
                return self.now - timedelta(days=days)
            else:
                return datetime.strptime(date_str, '%Y-%m-%d %H:%M')
        except (ValueError, AttributeError):
            try:
                return datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                return None

    def _format_timedelta(self, td: timedelta) -> str:
        days = td.days
        hours, remainder = divmod(td.seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        if days > 0: return f"{days}天 {hours}小时"
        elif hours > 0: return f"{hours}小时 {minutes}分钟"
        else: return f"{minutes}分钟"

    def _get_color_from_value(self, value, min_val, max_val, start_color, end_color):
        if max_val == min_val: return ''.join(f'{c:02x}' for c in start_color)
        percentage = (value - min_val) / (max_val - min_val)
        r = int(start_color[0] + (end_color[0] - start_color[0]) * percentage)
        g = int(start_color[1] + (end_color[1] - start_color[1]) * percentage)
        b = int(start_color[2] + (end_color[2] - start_color[2]) * percentage)
        return f"{r:02x}{g:02x}{b:02x}"

    def _parse_page_with_soup(self):
        self._log(f"开始解析页面: {self.driver.current_url}")
        try:
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.c_threadlist ul")))
            html_content = self.driver.page_source
            soup = BeautifulSoup(html_content, 'lxml')
            threads = soup.select("div.c_threadlist ul li")
            if not threads: self._log("警告：当前页面未找到任何主题帖。"); return
            self._log(f"在当前页面找到 {len(threads)} 个主题帖，正在高速解析...")

            for thread in threads:
                try:
                    title_tag = thread.select_one('div.tit > a[href^="thread-"]')
                    title = title_tag.text.strip() if title_tag else "N/A"
                    relative_url = title_tag.get('href') if title_tag else ""
                    thread_url = urljoin(self.base_url, relative_url)
                    date_tag = thread.select_one('div.dte')
                    creation_date_str = date_tag.text.strip() if date_tag else "N/A"
                    views_tag = thread.select_one('em.vie')
                    views_text = views_tag.text if views_tag else '0'
                    views = int("".join(filter(str.isdigit, views_text)))
                    last_reply_date_str = "N/A"
                    last_reply_tag_with_span = thread.select_one('em.y a[href*="goto=lastpost"] span[title]')
                    if last_reply_tag_with_span: last_reply_date_str = last_reply_tag_with_span.get('title', '').strip()
                    else:
                        last_reply_tag_direct = thread.select_one('em.y a[href*="goto=lastpost"]')
                        if last_reply_tag_direct: last_reply_date_str = last_reply_tag_direct.text.strip()
                    
                    daily_views = 0.0
                    time_since_publication_str, time_since_publication_days = "无法解析", float('inf')
                    creation_date = self._parse_human_readable_date(creation_date_str)
                    if creation_date:
                        days_diff = max((self.now - creation_date).days, 1)
                        daily_views = round(views / days_diff, 2)
                        time_since_publication_str = self._format_timedelta(self.now - creation_date)
                        time_since_publication_days = (self.now - creation_date).total_seconds() / 86400

                    time_since_last_reply_str, time_since_last_reply_hours = "无法解析", float('inf')
                    last_reply_datetime = self._parse_human_readable_date(last_reply_date_str)
                    if last_reply_datetime:
                        time_diff = self.now - last_reply_datetime
                        time_since_last_reply_str = self._format_timedelta(time_diff)
                        time_since_last_reply_hours = time_diff.total_seconds() / 3600

                    self.data.append({
                        "主题名称": title, "主题浏览量": views, "主题每日浏览量": daily_views,
                        "主题发布日期": creation_date_str, "主题发布距今时间": time_since_publication_str,
                        "主题发布距今天数": time_since_publication_days,
                        "主题最后回复日期": last_reply_date_str, "最后回复距今时间": time_since_last_reply_str,
                        "最后回复距今小时数": time_since_last_reply_hours,
                        "主题链接": thread_url
                    })
                except Exception as e: self._log(f"解析单个主题时发生未知错误，已跳过。错误: {e}")
        except TimeoutException: self._log("错误：页面加载超时或页面结构已改变，无法找到主题列表容器。"); self._take_screenshot(); raise

    def _navigate_to_next_page(self) -> bool:
        try:
            next_page_button = self.driver.find_element(By.CSS_SELECTOR, 'a.nxt')
            self._log("找到“下一页”按钮，正在导航...")
            self.driver.execute_script("arguments[0].click();", next_page_button)
            return True
        except NoSuchElementException:
            self._log("未找到“下一页”按钮，已到达最后一页。"); return False

    def _save_to_excel(self):
        if not self.data: self._log("没有抓取到任何数据，不生成Excel报告。"); return
        self._log(f"共抓取到 {len(self.data)} 条数据，正在生成Excel报告...")
        
        df = pd.DataFrame(self.data)
        df_sorted = df.sort_values(by="主题每日浏览量", ascending=False).reset_index(drop=True)
        
        columns_order = [
            "主题名称", "主题浏览量", "主题每日浏览量", "主题发布日期", "主题发布距今时间",
            "主题最后回复日期", "最后回复距今时间", "主题链接", 
            "主题发布距今天数", "最后回复距今小时数"
        ]
        df_final = df_sorted[columns_order]
        
        wb = Workbook(); ws = wb.active; ws.title = "主题帖子数据报告"

        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell_font = Font(name='微软雅黑', size=11)
        link_font = Font(name='微软雅黑', size=11, color="0563C1", underline='single')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
            ws.append(row)
            ws.row_dimensions[r_idx].height = 30
            for cell in ws[r_idx]:
                cell.border = border; cell.alignment = center_align
                if r_idx == 1:
                    cell.font = header_font; cell.fill = header_fill
                else:
                    cell.font = cell_font

        max_row = ws.max_row
        MINT_GREEN_RGB, CORAL_RED_RGB = (102, 187, 106), (239, 83, 80)
        
        views_rule = ColorScaleRule(start_type='min', start_color='FFF8E1', mid_type='percentile', mid_value=50, mid_color='FFB74D', end_type='max', end_color='81C784')
        ws.conditional_formatting.add(f"B2:C{max_row}", views_rule)

        for visible_col_name, helper_col_name, start_color, end_color in [("主题发布距今时间", "主题发布距今天数", MINT_GREEN_RGB, CORAL_RED_RGB), ("最后回复距今时间", "最后回复距今小时数", MINT_GREEN_RGB, CORAL_RED_RGB)]:
            visible_col_idx, helper_col_idx = df_final.columns.get_loc(visible_col_name) + 1, df_final.columns.get_loc(helper_col_name) + 1
            values = [v for v in df_final[helper_col_name] if isinstance(v, (int, float)) and v != float('inf')]
            if not values: continue
            min_val, max_val = min(values), max(values)
            for row_idx in range(2, max_row + 1):
                cell = ws.cell(row=row_idx, column=visible_col_idx)
                helper_val = ws.cell(row=row_idx, column=helper_col_idx).value
                if isinstance(helper_val, (int, float)) and helper_val != float('inf'):
                    cell.fill = PatternFill(start_color=self._get_color_from_value(helper_val, min_val, max_val, start_color, end_color), fill_type='solid')
        
        ws.column_dimensions[get_column_letter(df_final.columns.get_loc("主题发布距今天数") + 1)].hidden = True
        ws.column_dimensions[get_column_letter(df_final.columns.get_loc("最后回复距今小时数") + 1)].hidden = True

        url_col_idx = df_final.columns.get_loc("主题链接") + 1
        for row in range(2, max_row + 1):
            cell = ws.cell(row, url_col_idx)
            if cell.value and "http" in cell.value: cell.hyperlink = cell.value; cell.font = link_font
        
        column_widths = {'A': 100, 'B': 12, 'C': 15, 'D': 15, 'E': 18, 'F': 20, 'G': 18, 'H': 100}
        visible_columns = [col for col in df_final.columns if '距今天数' not in col and '距今小时数' not in col]
        for i, col_name in enumerate(visible_columns):
             ws.column_dimensions[get_column_letter(i+1)].width = list(column_widths.values())[i]

        ws.auto_filter.ref = f"A1:{get_column_letter(len(visible_columns))}{ws.max_row}"
        
        report_filename = f"苦力怕论坛帖子数据报告_{self.now.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        full_report_path = os.path.join(self.report_path, report_filename)
        wb.save(full_report_path)
        self._log(f"报告已成功保存至: {full_report_path}")

    def scrape(self, cookie_string: str):
        try:
            self._setup_driver()
            self._login_with_cookie(cookie_string)
            self._log(f"正在访问目标页面: {self.base_url}")
            self.driver.get(self.base_url)

            page_count = 1
            while True:
                self._log(f"--- 正在处理第 {page_count} 页 ---")
                self._parse_page_with_soup()
                if not self._navigate_to_next_page(): break
                page_count += 1
            self._save_to_excel()
        except Exception as e:
            self._log(f"在抓取过程中发生严重错误: {e}")
            self._take_screenshot()
        finally:
            if self.driver:
                self._log("关闭WebDriver。")
                self.driver.quit()
            self._log("爬取任务结束。")

if __name__ == '__main__':
    TARGET_URL = "https://klpbbs.com/home.php?mod=space&uid=2086093&do=thread&view=me&from=space"
    # 本地运行时，请确保此路径正确
    DRIVER_EXECUTABLE_PATH = os.path.join("chromedriver-win64", "chromedriver.exe")
    
    # 从环境变量中读取Cookie，这是为GitHub Actions设计的
    MY_COOKIE_STRING = os.getenv('KLPBBS_COOKIE', '')
    
    # 如果是在本地运行且环境变量为空，可以手动在此处填入用于调试
    if not MY_COOKIE_STRING:
        print("未在环境变量中找到Cookie，请在本地代码中手动填入Cookie字符串用于调试。")
        # MY_COOKIE_STRING = "在这里粘贴你的Cookie来进行本地调试"

    scraper = KlpbbsScraper(
        base_url=TARGET_URL,
        driver_path=DRIVER_EXECUTABLE_PATH
    )
    
    scraper.scrape(cookie_string=MY_COOKIE_STRING)
